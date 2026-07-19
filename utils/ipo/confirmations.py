"""Trade confirmations — consolidated Excel exports by buy/sell party."""

from __future__ import annotations

import io
import re
from typing import Any

from utils.ipo.categories import application_amount_from_ipo
from utils.ipo.service import list_positions


def _filters_match(
    pos: dict[str, Any],
    *,
    party_id: str | None,
    sell_needle: str,
    cat_needle: str,
    sub_needle: str,
) -> bool:
    if pos.get("is_premium"):
        return False
    if pos.get("allocation_status") != "Fully Allocated":
        return False
    if not pos.get("allocations"):
        return False
    if party_id and (pos.get("party_id") or "") != party_id:
        return False
    if cat_needle and (pos.get("category") or "") != cat_needle:
        return False
    if sub_needle and (pos.get("sub_category") or "") != sub_needle:
        return False
    if sell_needle:
        sells = list(pos.get("sells") or [])
        if not any(
            sell_needle in str(s.get("sell_party") or "").casefold() for s in sells
        ):
            return False
    return True


def _eligible_positions(
    broker_id: str,
    *,
    ipo_id: str | None = None,
    party_id: str | None = None,
    sell_party: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    category: str | None = None,
    sub_category: str | None = None,
) -> list[dict[str, Any]]:
    positions = list_positions(
        broker_id,
        ipo_id=ipo_id,
        party=None,
        date_from=date_from,
        date_to=date_to,
    )
    sell_needle = (sell_party or "").strip().casefold()
    cat_needle = (category or "").strip()
    sub_needle = (sub_category or "").strip()
    return [
        pos
        for pos in positions
        if _filters_match(
            pos,
            party_id=party_id,
            sell_needle=sell_needle,
            cat_needle=cat_needle,
            sub_needle=sub_needle,
        )
    ]


def list_confirmation_groups(
    broker_id: str,
    *,
    ipo_id: str | None = None,
    party_id: str | None = None,
    sell_party: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    category: str | None = None,
    sub_category: str | None = None,
) -> dict[str, list[dict[str, Any]]]:
    """
    Consolidate fully allocated trades into buy-party and sell-party groups
    so one Excel covers all matching trades for that counterpart.
    """
    positions = _eligible_positions(
        broker_id,
        ipo_id=ipo_id,
        party_id=party_id,
        sell_party=sell_party,
        date_from=date_from,
        date_to=date_to,
        category=category,
        sub_category=sub_category,
    )

    buy_map: dict[str, dict[str, Any]] = {}
    sell_map: dict[str, dict[str, Any]] = {}

    for pos in positions:
        ipo = pos.get("ipo") or {}
        ipo_name = ipo.get("display_name") or ipo.get("name") or ""
        buy_key = pos.get("party_id") or f"name:{(pos.get('party') or '').casefold()}"
        buy = buy_map.get(buy_key)
        if buy is None:
            buy = {
                "party_id": pos.get("party_id"),
                "party": pos.get("party") or "",
                "trade_count": 0,
                "applicant_count": 0,
                "ipo_names": set(),
                "categories": set(),
                "sub_categories": set(),
                "sell_parties": set(),
                "dates": set(),
            }
            buy_map[buy_key] = buy
        buy["trade_count"] += 1
        buy["applicant_count"] += int(pos.get("allocated_count") or 0)
        if ipo_name:
            buy["ipo_names"].add(ipo_name)
        if pos.get("category"):
            buy["categories"].add(pos["category"])
        if pos.get("sub_category"):
            buy["sub_categories"].add(pos["sub_category"])
        if pos.get("trade_date"):
            buy["dates"].add(str(pos["trade_date"])[:10])
        for s in pos.get("sells") or []:
            sp = (s.get("sell_party") or "").strip()
            if sp:
                buy["sell_parties"].add(sp)

        for sell in pos.get("sells") or []:
            sp = (sell.get("sell_party") or "").strip()
            if not sp:
                continue
            sell_key = sp.casefold()
            group = sell_map.get(sell_key)
            if group is None:
                group = {
                    "sell_party": sp,
                    "trade_count": 0,
                    "applicant_count": 0,
                    "ipo_names": set(),
                    "categories": set(),
                    "sub_categories": set(),
                    "buy_parties": set(),
                    "dates": set(),
                }
                sell_map[sell_key] = group
            group["trade_count"] += 1
            # Applicants on the position are what get confirmed to the sell party
            group["applicant_count"] += int(pos.get("allocated_count") or 0)
            if ipo_name:
                group["ipo_names"].add(ipo_name)
            if pos.get("category"):
                group["categories"].add(pos["category"])
            if pos.get("sub_category"):
                group["sub_categories"].add(pos["sub_category"])
            if pos.get("party"):
                group["buy_parties"].add(pos["party"])
            date_val = (sell.get("sell_date") or pos.get("trade_date") or "")[:10]
            if date_val:
                group["dates"].add(date_val)

    def _finalize_buy(item: dict[str, Any]) -> dict[str, Any]:
        dates = sorted(item["dates"])
        return {
            "party_id": item["party_id"],
            "party": item["party"],
            "trade_count": item["trade_count"],
            "applicant_count": item["applicant_count"],
            "ipo_names": sorted(item["ipo_names"]),
            "categories": sorted(item["categories"]),
            "sub_categories": sorted(item["sub_categories"]),
            "sell_parties": sorted(item["sell_parties"]),
            "date_from": dates[0] if dates else None,
            "date_to": dates[-1] if dates else None,
        }

    def _finalize_sell(item: dict[str, Any]) -> dict[str, Any]:
        dates = sorted(item["dates"])
        return {
            "sell_party": item["sell_party"],
            "trade_count": item["trade_count"],
            "applicant_count": item["applicant_count"],
            "ipo_names": sorted(item["ipo_names"]),
            "categories": sorted(item["categories"]),
            "sub_categories": sorted(item["sub_categories"]),
            "buy_parties": sorted(item["buy_parties"]),
            "date_from": dates[0] if dates else None,
            "date_to": dates[-1] if dates else None,
        }

    buy_parties = sorted(
        (_finalize_buy(v) for v in buy_map.values()),
        key=lambda x: (x["party"] or "").upper(),
    )
    sell_parties = sorted(
        (_finalize_sell(v) for v in sell_map.values()),
        key=lambda x: (x["sell_party"] or "").upper(),
    )
    return {"buy_parties": buy_parties, "sell_parties": sell_parties}


def build_client_confirmation_excel(
    broker_id: str,
    *,
    party_id: str,
    ipo_id: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    category: str | None = None,
    sub_category: str | None = None,
) -> tuple[bytes, str]:
    if not (party_id or "").strip():
        raise ValueError("Buy party is required for client confirmation.")
    positions = _eligible_positions(
        broker_id,
        ipo_id=ipo_id,
        party_id=party_id,
        date_from=date_from,
        date_to=date_to,
        category=category,
        sub_category=sub_category,
    )
    if not positions:
        raise ValueError("No fully allocated trades found for this buy party.")

    rows: list[dict[str, Any]] = []
    party_name = positions[0].get("party") or ""
    for pos in positions:
        ipo = pos.get("ipo") or {}
        sub = pos.get("sub_category") or ""
        app_amount = application_amount_from_ipo(ipo, sub)
        rate = float(pos.get("buy_rate") or 0)
        for app in pos.get("allocations") or []:
            rows.append(
                {
                    "name": app.get("name") or "",
                    "dpid": app.get("dpid") or "",
                    "pan": app.get("pan") or "",
                    "category": pos.get("category") or "",
                    "sub_category": sub,
                    "application_amount": app_amount,
                    "rate": rate,
                    "ipo_name": ipo.get("display_name") or ipo.get("name") or "",
                }
            )

    meta = _header_meta(rows)
    xlsx = _render_workbook(
        title="CLIENT CONFIRMATION",
        party_label="Buy Party",
        party_name=party_name,
        rate_label="BUY RATE",
        rows=rows,
        **meta,
    )
    filename = (
        f"Client_Confirmation_{_slug(party_name) or 'BuyParty'}"
        f"_{_slug(meta['ipo_name']) or 'IPO'}.xlsx"
    )
    return xlsx, filename


def build_broker_confirmation_excel(
    broker_id: str,
    *,
    sell_party: str,
    ipo_id: str | None = None,
    party_id: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    category: str | None = None,
    sub_category: str | None = None,
) -> tuple[bytes, str]:
    target = (sell_party or "").strip()
    if not target:
        raise ValueError("Sell party is required for broker confirmation.")
    target_key = target.casefold()

    positions = _eligible_positions(
        broker_id,
        ipo_id=ipo_id,
        party_id=party_id,
        sell_party=target,
        date_from=date_from,
        date_to=date_to,
        category=category,
        sub_category=sub_category,
    )
    rows: list[dict[str, Any]] = []
    display_name = target
    for pos in positions:
        matching = [
            s
            for s in (pos.get("sells") or [])
            if (s.get("sell_party") or "").strip().casefold() == target_key
        ]
        if not matching:
            # filter used contains-match; prefer exact group name from UI
            matching = [
                s
                for s in (pos.get("sells") or [])
                if target_key in (s.get("sell_party") or "").casefold()
            ]
        if not matching:
            continue
        # One confirmation block per position for this sell party (use first matching sell rate)
        sell = matching[0]
        display_name = (sell.get("sell_party") or display_name).strip() or display_name
        ipo = pos.get("ipo") or {}
        sub = pos.get("sub_category") or ""
        app_amount = application_amount_from_ipo(ipo, sub)
        rate = float(sell.get("sell_rate") or 0)
        for app in pos.get("allocations") or []:
            rows.append(
                {
                    "name": app.get("name") or "",
                    "dpid": app.get("dpid") or "",
                    "pan": app.get("pan") or "",
                    "category": pos.get("category") or "",
                    "sub_category": sub,
                    "application_amount": app_amount,
                    "rate": rate,
                    "ipo_name": ipo.get("display_name") or ipo.get("name") or "",
                }
            )

    if not rows:
        raise ValueError("No fully allocated sold trades found for this sell party.")

    meta = _header_meta(rows)
    xlsx = _render_workbook(
        title="BROKER CONFIRMATION",
        party_label="Sell Party",
        party_name=display_name,
        rate_label="SELL RATE",
        rows=rows,
        **meta,
    )
    filename = (
        f"Broker_Confirmation_{_slug(display_name) or 'SellParty'}"
        f"_{_slug(meta['ipo_name']) or 'IPO'}.xlsx"
    )
    return xlsx, filename


def _header_meta(rows: list[dict[str, Any]]) -> dict[str, str]:
    ipos = sorted({r.get("ipo_name") or "" for r in rows if r.get("ipo_name")})
    cats = sorted({r.get("category") or "" for r in rows if r.get("category")})
    subs = sorted({r.get("sub_category") or "" for r in rows if r.get("sub_category")})
    return {
        "ipo_name": ipos[0] if len(ipos) == 1 else ("Multiple IPOs" if ipos else ""),
        "category": cats[0] if len(cats) == 1 else ("Multiple" if cats else ""),
        "sub_category": subs[0] if len(subs) == 1 else ("Multiple" if subs else ""),
    }


def _slug(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", (value or "").strip())
    return cleaned.strip("_")[:40]


def _amount_cell(value: float | None) -> tuple[int | float | str, str]:
    """Whole numbers as integers; keep 2 decimals only when needed."""
    if value is None:
        return "", "#,##0"
    amount = float(value)
    if abs(amount - round(amount)) < 1e-9:
        return int(round(amount)), "#,##0"
    return amount, "#,##0.00"


def _render_workbook(
    *,
    title: str,
    ipo_name: str,
    party_label: str,
    party_name: str,
    category: str,
    sub_category: str,
    rate_label: str,
    rows: list[dict[str, Any]],
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for confirmation Excel export. Add openpyxl to requirements."
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Confirmation"

    header_font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    meta_label_font = Font(name="Calibri", bold=True, size=11)
    meta_value_font = Font(name="Calibri", size=11)
    col_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    cell_font = Font(name="Calibri", size=11)
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    title_fill = PatternFill("solid", fgColor="064E3B")
    col_fill = PatternFill("solid", fgColor="047857")
    alt_fill = PatternFill("solid", fgColor="ECFDF5")
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = header_font
    title_cell.fill = title_fill
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    ws["A3"] = "IPO"
    ws["B3"] = ipo_name
    ws["A4"] = party_label
    ws["B4"] = party_name
    ws["D3"] = "Category"
    ws["E3"] = category
    ws["D4"] = "Sub-Category"
    ws["E4"] = sub_category

    for row in (3, 4):
        for col in (1, 2, 4, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = meta_label_font if col in (1, 4) else meta_value_font
            cell.alignment = center

    headers = [
        "SR NO",
        "Applicant Name",
        "DPID",
        "PAN",
        "Category",
        "Application Amount",
        rate_label,
    ]
    header_row = 6
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = col_font
        cell.fill = col_fill
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[header_row].height = 22

    for idx, app in enumerate(rows, start=1):
        row = header_row + idx
        amount_value, amount_format = _amount_cell(app.get("application_amount"))
        values = [
            idx,
            app.get("name") or "",
            app.get("dpid") or "",
            app.get("pan") or "",
            app.get("category") or "",
            amount_value,
            float(app.get("rate") or 0),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value if value != "" else None)
            cell.font = cell_font
            cell.border = thin
            cell.alignment = center
            if idx % 2 == 0:
                cell.fill = alt_fill
            if col == 6 and value not in (None, ""):
                cell.number_format = amount_format
            elif col == 7 and value not in (None, ""):
                cell.number_format = "#,##0.00"

    widths = [8, 28, 18, 14, 18, 18, 14]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A7"
    ws.print_title_rows = "1:6"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
