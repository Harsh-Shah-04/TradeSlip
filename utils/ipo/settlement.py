"""Phase 3 — Settlement preview, draft, finalize, Excel export."""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import Any

import httpx
from openpyxl import Workbook

from utils.ipo.allotments import list_allotments_raw
from utils.ipo.categories import application_amount_from_ipo, is_premium, is_subject2
from utils.ipo.ledger import post_settlement_entries
from utils.ipo.models import (
    settlement_line_to_json,
    settlement_to_json,
)
from utils.ipo.service import get_ipo
from utils.supabase_client import _service_headers, _supabase_url

SETTLEMENTS_TABLE = "ipo_settlements"
LINES_TABLE = "ipo_settlement_lines"
HTTP_TIMEOUT = 30.0
MONEY_QUANT = Decimal("0.0001")

_http_client: httpx.Client | None = None


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _http() -> httpx.Client:
    global _http_client
    if _http_client is None:
        _http_client = httpx.Client(timeout=HTTP_TIMEOUT)
    return _http_client


def _money(value: float | int | Decimal | str | None) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    return Decimal(str(value)).quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)


def _direction(net_pl: float) -> str:
    if abs(net_pl) < 1e-9:
        return "Settled"
    return "Receivable" if net_pl > 0 else "Payable"


def _sell_side_direction(net_pl: float) -> str:
    """Sell parties (Ambica, Mama, …): positive listing P/L means you give to them."""
    if abs(net_pl) < 1e-9:
        return "Settled"
    return "Payable" if net_pl > 0 else "Receivable"


def _sell_side_financials(
    *,
    listing_sell_amt: float,
    seller_vyaj: float,
    buyer_vyaj: float,
    recorded_brokerage: float | None,
    premium: bool,
    subject2: bool = False,
    client_amount: float = 0.0,
) -> dict[str, float]:
    """Return sell-party settlement amounts without mixing buyer cost and brokerage."""
    if subject2:
        # Subject 2 uses three independent prices — never mixed:
        #   Buy Rate   → Guaranteed = Shares × Buy Rate (client guarantee, buy side)
        #   Sell Rate  → Contract = Shares × Sell Rate (seller_vyaj here) — the
        #                broker's receivable from the sell party, IN FULL
        #   Sold Price → Market = Shares × Allotment Sold Price — used only for the
        #                client leg: Difference = Guaranteed − Market
        # Sell side follows the same table math as other rows:
        #   net_pl = Market − Contract (negative ⇒ Receivable from the sell party)
        # i.e. the sell party owes the contract minus what the market sale already
        # realized. The client difference stays in settlement_difference for context.
        # Brokerage = Contract − Guaranteed = Shares × (Sell Rate − Buy Rate):
        # what we collect from the sell party minus what we pay the client.
        market = _money(listing_sell_amt)
        guaranteed = _money(client_amount)
        contract = _money(seller_vyaj)
        difference = guaranteed - market
        brokerage = float(contract - guaranteed)
        return {
            "sell_amt": float(market),
            "market_amount": float(market),
            "vyaj": float(contract),
            "settlement_difference": float(difference),
            "brokerage": brokerage,
            "profit": brokerage,
            "loss": 0.0,
            "net_pl": float(market - contract),
        }
    brokerage = (
        float(_money(recorded_brokerage))
        if recorded_brokerage is not None
        else float(_money(seller_vyaj) - _money(buyer_vyaj))
    )
    if premium:
        # Premium has no listing-sale leg: the sell party owes the full trade amount.
        return {
            "sell_amt": 0.0,
            "market_amount": float(_money(seller_vyaj)),
            "vyaj": float(_money(seller_vyaj)),
            "settlement_difference": brokerage,
            "brokerage": brokerage,
            "profit": max(brokerage, 0.0),
            "loss": max(-brokerage, 0.0),
            "net_pl": float(-_money(seller_vyaj)),
        }
    return {
        "sell_amt": float(_money(listing_sell_amt)),
        "market_amount": float(_money(listing_sell_amt)),
        "vyaj": float(_money(seller_vyaj)),
        "settlement_difference": brokerage,
        "brokerage": brokerage,
        "profit": max(brokerage, 0.0),
        "loss": max(-brokerage, 0.0),
        "net_pl": float(_money(listing_sell_amt) - _money(seller_vyaj)),
    }


def _positions_map(ipo_id: str) -> dict[str, dict[str, Any]]:
    """Position id → category / rates, used to apply category-specific settlement rules."""
    response = _http().get(
        f"{_supabase_url()}/rest/v1/ipo_positions",
        headers=_service_headers(),
        params={
            "select": "id,party,category,buy_app,buy_rate,buy_amt",
            "ipo_id": f"eq.{ipo_id}",
            "limit": "5000",
        },
        timeout=HTTP_TIMEOUT,
    )
    if response.status_code != 200:
        return {}
    return {str(r["id"]): r for r in (response.json() or []) if r.get("id")}


def _build_line_from_allotment(
    row: dict[str, Any],
    ipo: dict[str, Any],
    positions: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    applicant = row.get("ipo_applicants") if isinstance(row.get("ipo_applicants"), dict) else {}
    party = applicant.get("ipo_parties") if isinstance(applicant.get("ipo_parties"), dict) else {}
    status = row.get("status") or "Pending"
    shares = int(row.get("shares_allotted") or 0) if status == "Allotted" else 0
    allotted_apps = 1.0 if status == "Allotted" else 0.0
    position = (positions or {}).get(str(row.get("position_id") or "")) or {}
    subject2 = is_subject2(position.get("category"))
    s2_pending = False
    if subject2:
        # Subject 2 (guaranteed deal): the client sells the allotted shares at market
        # and keeps those proceeds. The market leg ALWAYS comes from this allotment
        # row's actual sold price — never from Position / Sell Trade sell rates.
        # Only Market − Guaranteed settles with the client:
        #   net > 0 → client returns the excess (Receivable)
        #   net < 0 → client gets topped up to the guarantee (Payable)
        # Nothing settles without an allotment.
        buy_rate = float(_money(position.get("buy_rate")))
        market_rate = None
        if row.get("is_sold") and row.get("sold_price") is not None:
            market_rate = float(_money(row.get("sold_price")))
        guaranteed = float(_money(shares) * _money(buy_rate)) if shares > 0 else 0.0
        if shares <= 0:
            vyaj = sell_amt = net_pl = 0.0
            sell_premium = 0.0
        elif market_rate is None:
            # Market sale not recorded yet — settlement difference unknown.
            s2_pending = True
            vyaj = guaranteed
            sell_amt = 0.0
            sell_premium = 0.0
            net_pl = 0.0
        else:
            vyaj = guaranteed
            sell_premium = float(_money(market_rate))
            sell_amt = float(_money(shares) * _money(market_rate))
            net_pl = float(_money(sell_amt) - _money(guaranteed))
    else:
        vyaj = float(_money(row.get("cost_per_app")))
        sell_premium = float(_money(row.get("sold_price"))) if row.get("is_sold") else 0.0
        sell_amt = float(_money(shares) * _money(sell_premium))
        net_pl = float(_money(sell_amt) - _money(vyaj))
    sub = row.get("sub_category") or ""
    return {
        "allotment_id": str(row.get("id") or ""),
        "party_id": str(applicant.get("party_id") or party.get("id") or "") or None,
        "applicant_id": str(row.get("applicant_id") or "") or None,
        "position_id": None if row.get("position_id") is None else str(row.get("position_id")),
        "party_name": party.get("name") or "",
        "applicant_name": applicant.get("name") or "",
        "pan": applicant.get("pan") or "",
        "dpid": applicant.get("dpid") or "",
        "sub_category": sub,
        "application_amount": application_amount_from_ipo(ipo, sub),
        "vyaj": vyaj,
        "applied": 1.0,
        "allotted_apps": allotted_apps,
        "shares_allotted": shares,
        "sell_premium": sell_premium,
        "sell_amt": sell_amt,
        "net_pl": net_pl,
        "direction": _direction(net_pl),
        "status": status,
        "is_sold": bool(row.get("is_sold")),
        "is_premium": False,
        "is_subject2": subject2,
        "s2_pending": s2_pending,
        "line_type": "Subject 2" if subject2 else "Application",
    }


def _allotments_by_position(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    by_pos: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        pid = row.get("position_id")
        if pid:
            by_pos[str(pid)].append(row)
    return by_pos


def _attribute_allotment_rows(
    sell_app: float, buy_app: float, allot_rows: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    """Map a grey-market sell qty to allotment rows on that position (FIFO by applicant name)."""
    if not allot_rows:
        return []
    ordered = sorted(
        allot_rows,
        key=lambda r: (
            ((r.get("ipo_applicants") or {}) if isinstance(r.get("ipo_applicants"), dict) else {}).get(
                "name"
            )
            or "",
            str(r.get("applicant_id") or ""),
        ),
    )
    if sell_app >= float(buy_app) - 1e-9:
        return ordered
    cap = max(0, min(int(round(sell_app)), len(ordered)))
    return ordered[:cap]


def _allotment_line_totals(rows: list[dict[str, Any]]) -> dict[str, float | int]:
    allotted = [r for r in rows if (r.get("status") or "") == "Allotted"]
    shares = sum(int(r.get("shares_allotted") or 0) for r in allotted)
    sell_amt = 0.0
    for row in allotted:
        if row.get("is_sold"):
            sell_amt += float(
                _money(row.get("shares_allotted") or 0) * _money(row.get("sold_price"))
            )
    vyaj = sum(float(_money(row.get("cost_per_app"))) for row in rows)
    return {
        "allotted": len(allotted),
        "shares_allotted": shares,
        "sell_amt": sell_amt,
        "vyaj": vyaj,
        "net_pl": float(_money(sell_amt) - _money(vyaj)),
    }


def _subject2_allotment_totals(
    rows: list[dict[str, Any]], buy_rate: float
) -> dict[str, float | int]:
    """Subject 2 guaranteed/market legs from attributed allotment rows.

    Both legs come from the SAME allotment rows: guaranteed = shares × position buy
    rate, market = shares × the row's actual sold price. Rows not yet marked sold
    contribute nothing — the difference is unknown until the sale price is recorded.
    Position / Sell Trade sell rates are never used here.
    """
    allotted = [r for r in rows if (r.get("status") or "") == "Allotted"]
    shares = sum(int(r.get("shares_allotted") or 0) for r in allotted)
    guaranteed = Decimal("0")
    market = Decimal("0")
    for row in allotted:
        if row.get("is_sold") and row.get("sold_price") is not None:
            qty = _money(row.get("shares_allotted") or 0)
            guaranteed += qty * _money(buy_rate)
            market += qty * _money(row.get("sold_price"))
    return {
        "allotted": len(allotted),
        "shares_allotted": shares,
        "guaranteed": float(guaranteed),
        "market": float(market),
    }


def _sell_party_summary_bucket(*, sell_party: str = "", line_type: str = "Application") -> dict[str, Any]:
    label = (sell_party or "—").strip() or "—"
    if line_type != "Application" and not label.endswith(f"({line_type})"):
        display = f"{label} ({line_type})"
    else:
        display = label
    return {
        "sell_party": display,
        "sell_party_base": label,
        "is_premium": line_type == "Premium",
        "is_subject2": line_type == "Subject 2",
        "line_type": line_type,
        "applied": 0.0,
        "allotted": 0.0,
        "shares_allotted": 0,
        "sell_amt": 0.0,
        "market_amount": 0.0,
        "vyaj": 0.0,
        "settlement_difference": 0.0,
        "brokerage": 0.0,
        "profit": 0.0,
        "loss": 0.0,
        "net_pl": 0.0,
    }


def _party_ids_by_name(names: list[str]) -> dict[str, str]:
    """Resolve Client Master party ids for position party names (best effort)."""
    cleaned = sorted({n.strip() for n in names if n and n.strip()})
    if not cleaned:
        return {}
    quoted = ",".join('"' + n.replace('"', '""') + '"' for n in cleaned)
    response = _http().get(
        f"{_supabase_url()}/rest/v1/ipo_parties",
        headers=_service_headers(),
        params={"select": "id,name", "name": f"in.({quoted})", "limit": "2000"},
        timeout=HTTP_TIMEOUT,
    )
    if response.status_code != 200:
        return {}
    out: dict[str, str] = {}
    for row in response.json() or []:
        name = str(row.get("name") or "").strip()
        if name and row.get("id"):
            out[name.casefold()] = str(row["id"])
    return out


def _premium_buy_lines(ipo_id: str) -> list[dict[str, Any]]:
    """Premium positions owe the buy party the full buy amount (no listing leg)."""
    response = _http().get(
        f"{_supabase_url()}/rest/v1/ipo_positions",
        headers=_service_headers(),
        params={
            "select": "id,party,category,buy_app,buy_amt",
            "ipo_id": f"eq.{ipo_id}",
            "limit": "5000",
        },
        timeout=HTTP_TIMEOUT,
    )
    if response.status_code != 200:
        return []
    premium_positions = [
        r for r in (response.json() or []) if is_premium(r.get("category"))
    ]
    if not premium_positions:
        return []
    party_ids = _party_ids_by_name([str(r.get("party") or "") for r in premium_positions])
    lines: list[dict[str, Any]] = []
    for pos in premium_positions:
        party = str(pos.get("party") or "").strip() or "—"
        buy_amt = float(_money(pos.get("buy_amt")))
        shares = int(round(float(pos.get("buy_app") or 0)))
        net_pl = float(-_money(buy_amt))
        lines.append(
            {
                "allotment_id": None,
                "party_id": party_ids.get(party.casefold()),
                "applicant_id": None,
                "position_id": str(pos.get("id") or "") or None,
                "party_name": party,
                "applicant_name": "Premium shares",
                "pan": "",
                "dpid": "",
                "sub_category": pos.get("category") or "Premium",
                "application_amount": None,
                "vyaj": buy_amt,
                # Premium is share-qty, not IPO applications — mirror sell-side Applied.
                "applied": float(shares),
                "allotted_apps": 0.0,
                "shares_allotted": shares,
                "sell_premium": 0.0,
                # No listing-sale leg with the buy party; grey-market sell is Ambica's side.
                "sell_amt": 0.0,
                "net_pl": net_pl,
                "direction": _direction(net_pl),
                "status": "Premium",
                "is_sold": True,
                "is_premium": True,
                "line_type": "Premium",
            }
        )
    lines.sort(key=lambda l: (l["party_name"].lower(), l["position_id"] or ""))
    return lines


def _party_summary(lines: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_party: dict[str, dict[str, Any]] = {}
    for line in lines:
        key = line.get("party_id") or line.get("party_name") or "unknown"
        bucket = by_party.setdefault(
            key,
            {
                "party_id": line.get("party_id"),
                "party_name": line.get("party_name") or "—",
                "applied": 0.0,
                "allotted": 0.0,
                "shares_allotted": 0,
                "sell_amt": 0.0,
                "vyaj": 0.0,
                "net_pl": 0.0,
                "line_types": set(),
            },
        )
        bucket["applied"] += float(line.get("applied") or 0)
        bucket["allotted"] += float(line.get("allotted_apps") or 0)
        bucket["shares_allotted"] += int(line.get("shares_allotted") or 0)
        bucket["sell_amt"] += float(line.get("sell_amt") or 0)
        bucket["vyaj"] += float(line.get("vyaj") or 0)
        bucket["net_pl"] += float(line.get("net_pl") or 0)
        line_type = line.get("line_type") or ("Premium" if line.get("is_premium") else "Application")
        bucket["line_types"].add(line_type)
    result = []
    for bucket in by_party.values():
        bucket["direction"] = _direction(bucket["net_pl"])
        # average sell premium for display when shares > 0
        if bucket["shares_allotted"] > 0 and bucket["sell_amt"]:
            bucket["sell_premium"] = bucket["sell_amt"] / bucket["shares_allotted"]
        else:
            bucket["sell_premium"] = 0.0
        types = bucket.pop("line_types")
        bucket["line_type"] = next(iter(types)) if len(types) == 1 else "Mixed"
        bucket["is_premium"] = bucket["line_type"] == "Premium"
        bucket["is_subject2"] = bucket["line_type"] == "Subject 2"
        result.append(bucket)
    result.sort(key=lambda x: (x.get("party_name") or "").lower())
    return result


def _grey_market_sell_side(ipo_id: str) -> dict[str, Any]:
    """Grey-market sells linked to allotments — Application and Premium are separate rows."""
    empty: dict[str, Any] = {
        "sell_party_summary": [],
        "sell_lines": [],
        "sell_party_totals": {},
        "grey_market_brokerage": 0.0,
    }
    pos_resp = _http().get(
        f"{_supabase_url()}/rest/v1/ipo_positions",
        headers=_service_headers(),
        params={
            "select": "id,party,buy_rate,buy_app,category",
            "ipo_id": f"eq.{ipo_id}",
            "limit": "5000",
        },
        timeout=HTTP_TIMEOUT,
    )
    if pos_resp.status_code != 200:
        return empty
    positions = {
        str(r["id"]): r for r in (pos_resp.json() or []) if r.get("id")
    }
    if not positions:
        return empty

    allot_by_pos = _allotments_by_position(list_allotments_raw(ipo_id))

    sell_lines: list[dict[str, Any]] = []
    ids = list(positions.keys())
    for i in range(0, len(ids), 100):
        chunk = ids[i : i + 100]
        id_filter = "(" + ",".join(chunk) + ")"
        sell_resp = _http().get(
            f"{_supabase_url()}/rest/v1/ipo_sells",
            headers=_service_headers(),
            params={
                "select": (
                    "id,position_id,sell_date,sell_app,sell_rate,"
                    "sell_amt,sell_party,brokerage"
                ),
                "position_id": f"in.{id_filter}",
                "order": "sell_date.asc",
                "limit": "5000",
            },
            timeout=HTTP_TIMEOUT,
        )
        if sell_resp.status_code != 200:
            continue
        for s in sell_resp.json() or []:
            pos = positions.get(str(s.get("position_id")) or "") or {}
            sell_qty = float(s.get("sell_app") or 0)
            buy_app = float(pos.get("buy_app") or 0)
            buy_rate = float(pos.get("buy_rate") or 0)
            pid = str(s.get("position_id") or "")
            sell_party = (s.get("sell_party") or "").strip() or "—"
            premium = is_premium(pos.get("category"))
            subject2 = is_subject2(pos.get("category"))
            line_type = "Premium" if premium else "Subject 2" if subject2 else "Application"
            grey_sell_amt = float(_money(s.get("sell_amt")))
            buyer_vyaj = float(_money(sell_qty) * _money(buy_rate))
            recorded_brokerage = s.get("brokerage")

            if premium:
                applied = sell_qty
                allotted = 0.0
                shares_allotted = int(round(sell_qty))
                financials = _sell_side_financials(
                    listing_sell_amt=0.0,
                    seller_vyaj=grey_sell_amt,
                    buyer_vyaj=buyer_vyaj,
                    recorded_brokerage=recorded_brokerage,
                    premium=True,
                )
            elif subject2:
                # Subject 2 settles on allotted shares, never on applications.
                # No allotment ⇒ deal never executed ⇒ everything stays 0.
                # Contract (vyaj) = allotted shares × this sell trade's sell_rate —
                # receivable from the sell party in full. Guaranteed/market come
                # from the attributed allotments (buy_rate / actual sold prices)
                # and only feed the client-difference context fields.
                attr_rows = _attribute_allotment_rows(
                    sell_qty, buy_app, allot_by_pos.get(pid, [])
                )
                s2_totals = _subject2_allotment_totals(attr_rows, buy_rate)
                applied = sell_qty
                allotted = float(s2_totals["allotted"])
                shares_allotted = int(s2_totals["shares_allotted"])
                market_amount = float(s2_totals["market"])
                guaranteed_amount = float(s2_totals["guaranteed"])
                sell_rate = float(s.get("sell_rate") or 0)
                contract_amount = float(_money(shares_allotted) * _money(sell_rate))
                buyer_vyaj = guaranteed_amount
                financials = _sell_side_financials(
                    listing_sell_amt=market_amount,
                    seller_vyaj=contract_amount,
                    buyer_vyaj=guaranteed_amount,
                    recorded_brokerage=None,
                    premium=False,
                    subject2=True,
                    client_amount=guaranteed_amount,
                )
            else:
                attr_rows = _attribute_allotment_rows(
                    sell_qty, buy_app, allot_by_pos.get(pid, [])
                )
                totals = _allotment_line_totals(attr_rows)
                applied = sell_qty
                allotted = float(totals["allotted"])
                shares_allotted = int(totals["shares_allotted"])
                financials = _sell_side_financials(
                    listing_sell_amt=float(totals["sell_amt"]),
                    seller_vyaj=grey_sell_amt,
                    buyer_vyaj=buyer_vyaj,
                    recorded_brokerage=recorded_brokerage,
                    premium=False,
                )

            sell_lines.append(
                {
                    "sell_id": str(s.get("id") or ""),
                    "sell_date": str(s.get("sell_date") or "")[:10],
                    "sell_party": sell_party,
                    "sell_party_display": (
                        f"{sell_party} ({line_type})" if line_type != "Application" else sell_party
                    ),
                    "buy_party": pos.get("party") or "—",
                    "is_premium": premium,
                    "is_subject2": subject2,
                    "line_type": line_type,
                    "applied": applied,
                    "allotted": allotted,
                    "shares_allotted": shares_allotted,
                    "sell_amt": financials["sell_amt"],
                    "market_amount": financials["market_amount"],
                    "vyaj": financials["vyaj"],
                    "buyer_vyaj": buyer_vyaj,
                    "settlement_difference": financials["settlement_difference"],
                    "brokerage": financials["brokerage"],
                    "profit": financials["profit"],
                    "loss": financials["loss"],
                    "net_pl": financials["net_pl"],
                    "direction": _sell_side_direction(financials["net_pl"]),
                    "sell_rate": float(s.get("sell_rate") or 0),
                    "grey_sell_amt": grey_sell_amt,
                }
            )

    # Key by sell party + line type so Application / Premium / Subject 2 never club together.
    by_party: dict[str, dict[str, Any]] = {}
    for line in sell_lines:
        base = line["sell_party"] or "—"
        line_type = line.get("line_type") or "Application"
        key = f"{base.casefold()}::{line_type}"
        bucket = by_party.setdefault(
            key,
            _sell_party_summary_bucket(sell_party=base, line_type=line_type),
        )
        bucket["applied"] += float(line["applied"])
        bucket["allotted"] += float(line["allotted"])
        bucket["shares_allotted"] += int(line["shares_allotted"])
        bucket["sell_amt"] += float(line["sell_amt"])
        bucket["market_amount"] += float(line["market_amount"])
        bucket["vyaj"] += float(line["vyaj"])
        bucket["settlement_difference"] += float(line["settlement_difference"])
        bucket["brokerage"] += float(line["brokerage"])
        bucket["profit"] += float(line["profit"])
        bucket["loss"] += float(line["loss"])
        bucket["net_pl"] += float(line["net_pl"])

    summary = []
    for bucket in by_party.values():
        bucket["direction"] = _sell_side_direction(bucket["net_pl"])
        summary.append(bucket)
    summary.sort(
        key=lambda x: (
            (x.get("sell_party_base") or "").lower(),
            x.get("line_type") or "",
        )
    )

    totals = {
        "applied": float(_money(sum(p["applied"] for p in summary))),
        "allotted": float(_money(sum(p["allotted"] for p in summary))),
        "shares_allotted": sum(int(p["shares_allotted"]) for p in summary),
        "sell_amt": float(_money(sum(p["sell_amt"] for p in summary))),
        "market_amount": float(_money(sum(p["market_amount"] for p in summary))),
        "vyaj": float(_money(sum(p["vyaj"] for p in summary))),
        "settlement_difference": float(
            _money(sum(p["settlement_difference"] for p in summary))
        ),
        "brokerage": float(_money(sum(p["brokerage"] for p in summary))),
        "profit": float(_money(sum(p["profit"] for p in summary))),
        "loss": float(_money(sum(p["loss"] for p in summary))),
        "net_pl": float(_money(sum(p["net_pl"] for p in summary))),
        "direction": _sell_side_direction(
            float(_money(sum(p["net_pl"] for p in summary)))
        ),
    }
    return {
        "sell_party_summary": summary,
        "sell_lines": sell_lines,
        "grey_market_brokerage": totals["brokerage"],
        "sell_party_totals": totals,
    }


def preview_settlement(ipo_id: str) -> dict[str, Any]:
    ipo = get_ipo(ipo_id, include_trade_count=False)
    rows = list_allotments_raw(ipo_id)
    positions = _positions_map(ipo_id)
    lines = [_build_line_from_allotment(r, ipo, positions) for r in rows]
    lines.extend(_premium_buy_lines(ipo_id))
    warnings: list[str] = []
    unsold = [l for l in lines if l["status"] == "Allotted" and not l["is_sold"]]
    if unsold:
        warnings.append(f"{len(unsold)} allotted applicant(s) not yet marked sold.")
    pending = [l for l in lines if l["status"] == "Pending"]
    if pending:
        warnings.append(f"{len(pending)} allotment(s) still Pending.")
    unlinked_premium = [
        l for l in lines if l.get("is_premium") and not l.get("party_id")
    ]
    if unlinked_premium:
        warnings.append(
            f"{len(unlinked_premium)} premium buy line(s) have no Client Master party — "
            "they show here but won't post to the ledger."
        )
    s2_pending = [l for l in lines if l.get("s2_pending")]
    if s2_pending:
        warnings.append(
            f"{len(s2_pending)} Subject 2 allotment(s) have no sold price recorded — "
            "mark them sold with the actual sold price (Allotments) to settle the difference."
        )
    party = _party_summary(lines)
    sell_side = _grey_market_sell_side(ipo_id)
    if not sell_side["sell_party_summary"]:
        warnings.append("No grey-market sells recorded yet (Ambica / Mama / sell parties).")
    return {
        "ipo": ipo,
        "lines": lines,
        "party_summary": party,
        "sell_party_summary": sell_side["sell_party_summary"],
        "sell_lines": sell_side["sell_lines"],
        "sell_party_totals": sell_side.get("sell_party_totals") or {},
        "grey_market_brokerage": sell_side["grey_market_brokerage"],
        "totals": {
            "applied": sum(l["applied"] for l in lines),
            "allotted": sum(l["allotted_apps"] for l in lines),
            "shares_allotted": sum(l["shares_allotted"] for l in lines),
            "sell_amt": sum(l["sell_amt"] for l in lines),
            "vyaj": sum(l["vyaj"] for l in lines),
            "net_pl": sum(l["net_pl"] for l in lines),
        },
        "warnings": warnings,
    }


def _get_settlement(settlement_id: str) -> dict[str, Any]:
    response = _http().get(
        f"{_supabase_url()}/rest/v1/{SETTLEMENTS_TABLE}",
        headers=_service_headers(),
        params={"id": f"eq.{settlement_id}", "select": "*", "limit": "1"},
        timeout=HTTP_TIMEOUT,
    )
    if response.status_code != 200:
        raise RuntimeError(f"Get settlement failed ({response.status_code})")
    rows = response.json()
    if not isinstance(rows, list) or not rows:
        raise LookupError("Settlement not found.")
    return rows[0]


def _list_lines(settlement_id: str) -> list[dict[str, Any]]:
    response = _http().get(
        f"{_supabase_url()}/rest/v1/{LINES_TABLE}",
        headers=_service_headers(),
        params={
            "settlement_id": f"eq.{settlement_id}",
            "select": "*",
            "order": "party_name.asc,applicant_name.asc",
            "limit": "20000",
        },
        timeout=HTTP_TIMEOUT,
    )
    if response.status_code != 200:
        raise RuntimeError(f"List settlement lines failed ({response.status_code})")
    rows = response.json()
    return rows if isinstance(rows, list) else []


def get_settlement(settlement_id: str) -> dict[str, Any]:
    row = _get_settlement(settlement_id)
    lines = [settlement_line_to_json(l) for l in _list_lines(settlement_id)]
    # Stored lines don't persist line_type — re-derive it from the position category.
    positions = _positions_map(str(row["ipo_id"]))
    for line in lines:
        pos = positions.get(str(line.get("position_id") or "")) or {}
        category = pos.get("category")
        if is_subject2(category):
            line["line_type"] = "Subject 2"
            line["is_subject2"] = True
        elif is_premium(category):
            line["line_type"] = "Premium"
            line["is_premium"] = True
        else:
            line["line_type"] = "Application"
    out = settlement_to_json(row, lines=lines)
    out["party_summary"] = _party_summary(lines)
    out["ipo"] = get_ipo(str(row["ipo_id"]), include_trade_count=False)
    sell_side = _grey_market_sell_side(str(row["ipo_id"]))
    out["sell_party_summary"] = sell_side["sell_party_summary"]
    out["sell_lines"] = sell_side["sell_lines"]
    out["sell_party_totals"] = sell_side.get("sell_party_totals") or {}
    out["grey_market_brokerage"] = sell_side["grey_market_brokerage"]
    return out


def create_or_refresh_draft(ipo_id: str, broker_id: str, notes: str = "") -> dict[str, Any]:
    # Reuse existing draft for this IPO if present
    existing = _http().get(
        f"{_supabase_url()}/rest/v1/{SETTLEMENTS_TABLE}",
        headers=_service_headers(),
        params={
            "ipo_id": f"eq.{ipo_id}",
            "status": "eq.Draft",
            "select": "*",
            "limit": "1",
        },
        timeout=HTTP_TIMEOUT,
    )
    preview = preview_settlement(ipo_id)
    ipo = preview["ipo"]
    draft_id = None
    if existing.status_code == 200 and isinstance(existing.json(), list) and existing.json():
        draft_id = str(existing.json()[0]["id"])
        # wipe lines
        _http().delete(
            f"{_supabase_url()}/rest/v1/{LINES_TABLE}",
            headers=_service_headers({"Prefer": "return=minimal"}),
            params={"settlement_id": f"eq.{draft_id}"},
            timeout=HTTP_TIMEOUT,
        )
        _http().patch(
            f"{_supabase_url()}/rest/v1/{SETTLEMENTS_TABLE}",
            headers=_service_headers({"Content-Type": "application/json"}),
            params={"id": f"eq.{draft_id}"},
            json={
                "listing_price_used": ipo.get("listing_price"),
                "notes": notes or "",
                "updated_at": _now(),
            },
            timeout=HTTP_TIMEOUT,
        )
    else:
        # Block if already finalized
        fin = _http().get(
            f"{_supabase_url()}/rest/v1/{SETTLEMENTS_TABLE}",
            headers=_service_headers(),
            params={
                "ipo_id": f"eq.{ipo_id}",
                "status": "eq.Finalized",
                "select": "id",
                "limit": "1",
            },
            timeout=HTTP_TIMEOUT,
        )
        if fin.status_code == 200 and isinstance(fin.json(), list) and fin.json():
            raise ValueError("This IPO already has a finalized settlement.")
        create = _http().post(
            f"{_supabase_url()}/rest/v1/{SETTLEMENTS_TABLE}",
            headers=_service_headers(
                {"Content-Type": "application/json", "Prefer": "return=representation"}
            ),
            json={
                "ipo_id": ipo_id,
                "broker_id": broker_id,
                "status": "Draft",
                "listing_price_used": ipo.get("listing_price"),
                "notes": notes or "",
                "updated_at": _now(),
            },
            timeout=HTTP_TIMEOUT,
        )
        if create.status_code not in (200, 201):
            raise RuntimeError(f"Create settlement failed ({create.status_code}): {create.text[:300]}")
        data = create.json()
        created = data[0] if isinstance(data, list) else data
        draft_id = str(created["id"])

    line_rows = []
    for line in preview["lines"]:
        line_rows.append(
            {
                "settlement_id": draft_id,
                "allotment_id": line.get("allotment_id"),
                "party_id": line.get("party_id"),
                "applicant_id": line.get("applicant_id"),
                "position_id": line.get("position_id"),
                "party_name": line.get("party_name") or "",
                "applicant_name": line.get("applicant_name") or "",
                "pan": line.get("pan") or "",
                "dpid": line.get("dpid") or "",
                "sub_category": line.get("sub_category") or "",
                "application_amount": line.get("application_amount"),
                "vyaj": line["vyaj"],
                "applied": line["applied"],
                "allotted_apps": line["allotted_apps"],
                "shares_allotted": line["shares_allotted"],
                "sell_premium": line["sell_premium"],
                "sell_amt": line["sell_amt"],
                "net_pl": line["net_pl"],
                "direction": line["direction"],
            }
        )
    for i in range(0, len(line_rows), 100):
        chunk = line_rows[i : i + 100]
        if not chunk:
            continue
        resp = _http().post(
            f"{_supabase_url()}/rest/v1/{LINES_TABLE}",
            headers=_service_headers(
                {"Content-Type": "application/json", "Prefer": "return=minimal"}
            ),
            json=chunk,
            timeout=HTTP_TIMEOUT,
        )
        if resp.status_code not in (200, 201):
            raise RuntimeError(f"Insert settlement lines failed ({resp.status_code}): {resp.text[:300]}")

    out = get_settlement(draft_id)
    out["warnings"] = preview.get("warnings") or []
    out["grey_market_brokerage"] = preview.get("grey_market_brokerage")
    out["sell_party_summary"] = preview.get("sell_party_summary") or []
    out["sell_lines"] = preview.get("sell_lines") or []
    out["sell_party_totals"] = preview.get("sell_party_totals") or {}
    return out


def finalize_settlement(settlement_id: str) -> dict[str, Any]:
    row = _get_settlement(settlement_id)
    if row.get("status") == "Finalized":
        return get_settlement(settlement_id)
    lines = _list_lines(settlement_id)
    if not lines:
        raise ValueError("Settlement has no lines. Save a draft first.")

    patch = _http().patch(
        f"{_supabase_url()}/rest/v1/{SETTLEMENTS_TABLE}",
        headers=_service_headers({"Content-Type": "application/json"}),
        params={"id": f"eq.{settlement_id}", "status": "eq.Draft"},
        json={
            "status": "Finalized",
            "finalized_at": _now(),
            "updated_at": _now(),
        },
        timeout=HTTP_TIMEOUT,
    )
    if patch.status_code not in (200, 204):
        raise RuntimeError(f"Finalize failed ({patch.status_code}): {patch.text[:300]}")

    # Post ledger: one entry per party for this settlement
    party_nets: dict[str, float] = defaultdict(float)
    party_names: dict[str, str] = {}
    for line in lines:
        pid = line.get("party_id")
        if not pid:
            continue
        party_nets[str(pid)] += float(line.get("net_pl") or 0)
        party_names[str(pid)] = line.get("party_name") or ""

    post_settlement_entries(
        ipo_id=str(row["ipo_id"]),
        settlement_id=settlement_id,
        party_nets=dict(party_nets),
    )
    return get_settlement(settlement_id)


def build_settlement_excel(settlement_id: str) -> bytes:
    data = get_settlement(settlement_id)
    lines = data.get("lines") or []
    party = data.get("party_summary") or _party_summary(lines)
    sell_parties = data.get("sell_party_summary") or []
    sell_lines = data.get("sell_lines") or []
    ipo = data.get("ipo") or {}
    title = ipo.get("display_name") or ipo.get("name") or "IPO"

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Party Summary"
    ws1.append(
        [
            "summary",
            "applied",
            "alloted",
            "allote share",
            "sell premium",
            "sellamt",
            "vyaj",
            "net p/l",
            "direction",
        ]
    )
    for p in party:
        ws1.append(
            [
                p.get("party_name"),
                p.get("applied"),
                p.get("allotted"),
                p.get("shares_allotted"),
                p.get("sell_premium"),
                p.get("sell_amt"),
                p.get("vyaj"),
                p.get("net_pl"),
                p.get("direction"),
            ]
        )

    ws2 = wb.create_sheet("Applicants")
    ws2.append(
        [
            "SR NO",
            "NAME",
            "DPID",
            "PAN",
            "CATEGORY",
            "AMT",
            "VYAJ",
            "applied",
            "allotted",
            "shares",
            "sell premium",
            "sellamt",
            "net p/l",
            "direction",
        ]
    )
    for i, line in enumerate(lines, start=1):
        ws2.append(
            [
                i,
                line.get("applicant_name"),
                line.get("dpid"),
                line.get("pan"),
                line.get("sub_category"),
                line.get("application_amount"),
                line.get("vyaj"),
                line.get("applied"),
                line.get("allotted_apps"),
                line.get("shares_allotted"),
                line.get("sell_premium"),
                line.get("sell_amt"),
                line.get("net_pl"),
                line.get("direction"),
            ]
        )

    ws3 = wb.create_sheet("Sell Party Summary")
    ws3.append(
        [
            "SELL PARTY",
            "TYPE",
            "APPLIED",
            "ALLOTTED",
            "SHARES",
            "MARKET / LISTING AMOUNT",
            "VYAJ / CONTRACT AMOUNT",
            "BROKERAGE",
            "SETTLEMENT DIFFERENCE",
            "RECEIVABLE / PAYABLE",
            "DIRECTION",
        ]
    )
    for row in sell_parties:
        ws3.append(
            [
                row.get("sell_party_base") or row.get("sell_party"),
                row.get("line_type"),
                row.get("applied"),
                row.get("allotted"),
                row.get("shares_allotted"),
                row.get("market_amount", row.get("sell_amt")),
                row.get("vyaj"),
                row.get("brokerage"),
                row.get("settlement_difference", row.get("brokerage")),
                row.get("net_pl"),
                row.get("direction"),
            ]
        )

    ws4 = wb.create_sheet("Sell Trades")
    ws4.append(
        [
            "DATE",
            "SELL PARTY",
            "TYPE",
            "BUY PARTY",
            "APPLIED",
            "ALLOTTED",
            "SHARES",
            "MARKET / LISTING AMOUNT",
            "VYAJ / CONTRACT AMOUNT",
            "CLIENT AMOUNT",
            "BROKERAGE",
            "SETTLEMENT DIFFERENCE",
            "RECEIVABLE / PAYABLE",
            "DIRECTION",
        ]
    )
    for row in sell_lines:
        ws4.append(
            [
                row.get("sell_date"),
                row.get("sell_party"),
                row.get("line_type"),
                row.get("buy_party"),
                row.get("applied"),
                row.get("allotted"),
                row.get("shares_allotted"),
                row.get("market_amount", row.get("sell_amt")),
                row.get("vyaj"),
                row.get("buyer_vyaj"),
                row.get("brokerage"),
                row.get("settlement_difference", row.get("brokerage")),
                row.get("net_pl"),
                row.get("direction"),
            ]
        )

    ws1.cell(1, 11, title)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
