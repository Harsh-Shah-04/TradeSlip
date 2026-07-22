"""Phase 3 — Continuous Client Ledger (open-item, buyers + sell parties).

Model
-----
Every ledger entry belongs to exactly one ACCOUNT, which is either a buy party
(`ipo_parties`) or a sell party (`ipo_sell_parties`).

Entries are of two kinds:

* **Charges** (`Settlement`, `Adjustment`) — one per IPO. These are *open items*:
  they stay Pending until payments are allocated against them.
* **Payments** (`PaymentReceived`, `PaymentPaid`) — allocated to specific charges
  through `ipo_ledger_allocations`. A payment may also sit unallocated ("on
  account").

Sign convention on `amount` is unchanged: **+ collect** (money coming in),
**− pay them** (money going out). Allocation amounts are always positive magnitudes;
the direction comes from the charge they clear.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import Any

import httpx
from openpyxl import Workbook

from utils.ipo.models import ledger_entry_to_json
from utils.supabase_client import _service_headers, _supabase_url

LEDGER_TABLE = "ipo_ledger_entries"
ALLOCATIONS_TABLE = "ipo_ledger_allocations"
PARTIES_TABLE = "ipo_parties"
SELL_PARTIES_TABLE = "ipo_sell_parties"
IPO_TABLE = "ipo_master"
HTTP_TIMEOUT = 30.0
MONEY_QUANT = Decimal("0.0001")
EPSILON = 0.005  # sub-paisa noise is "settled"

ACCOUNT_PARTY = "party"
ACCOUNT_SELL_PARTY = "sell_party"
ACCOUNT_TYPES = (ACCOUNT_PARTY, ACCOUNT_SELL_PARTY)

CHARGE_TYPES = ("Settlement", "Adjustment")
PAYMENT_TYPES = ("PaymentReceived", "PaymentPaid")

_http_client: httpx.Client | None = None


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _http() -> httpx.Client:
    global _http_client
    if _http_client is None:
        _http_client = httpx.Client(timeout=HTTP_TIMEOUT)
    return _http_client


def _money(value: float | int | Decimal | str | None) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    return Decimal(str(value)).quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)


# --------------------------------------------------------------------- accounts


def _account_column(account_type: str) -> str:
    if account_type == ACCOUNT_PARTY:
        return "party_id"
    if account_type == ACCOUNT_SELL_PARTY:
        return "sell_party_id"
    raise ValueError(f"account_type must be one of {ACCOUNT_TYPES}")


def _account_table(account_type: str) -> str:
    return PARTIES_TABLE if account_type == ACCOUNT_PARTY else SELL_PARTIES_TABLE


def side_to_account_type(side: str | None) -> str:
    """UI 'buyers' / 'sellers' tab → account type."""
    key = (side or "").strip().casefold()
    if key in ("seller", "sellers", "sell_party", "sell"):
        return ACCOUNT_SELL_PARTY
    return ACCOUNT_PARTY


# ------------------------------------------------------------- pure math (tested)


def charge_status(amount: float, allocated: float) -> str:
    """Pending / Part paid / Done for a single charge."""
    outstanding = abs(float(amount)) - float(allocated)
    if abs(float(amount)) < EPSILON:
        return "Done"
    if outstanding <= EPSILON:
        return "Done"
    if float(allocated) > EPSILON:
        return "Part paid"
    return "Pending"


def charge_direction(amount: float) -> str:
    """Which way a charge moves money, in the trader's words."""
    if abs(float(amount)) < EPSILON:
        return "Cleared"
    return "Collect" if float(amount) > 0 else "Pay them"


def balance_direction(balance: float) -> str:
    if abs(float(balance)) < EPSILON:
        return "Cleared"
    return "Collect" if float(balance) > 0 else "Pay them"


def _outstanding(amount: float, allocated: float) -> float:
    return max(0.0, float(_money(abs(float(amount)) - float(allocated))))


def summarize_entries(
    entries: list[dict[str, Any]], allocated_by_entry: dict[str, float]
) -> dict[str, float | int]:
    """Roll a single account's entries up into headline numbers.

    Pure — no I/O — so the arithmetic can be unit tested.
    """
    balance = Decimal("0")
    to_receive = Decimal("0")
    to_pay = Decimal("0")
    on_account = Decimal("0")
    open_count = 0
    for entry in entries:
        amount = float(entry.get("amount") or 0)
        balance += _money(amount)
        entry_id = str(entry.get("id") or "")
        allocated = float(allocated_by_entry.get(entry_id, 0.0))
        outstanding = _outstanding(amount, allocated)
        if entry.get("entry_type") in CHARGE_TYPES:
            if outstanding > EPSILON:
                open_count += 1
                if amount > 0:
                    to_receive += _money(outstanding)
                else:
                    to_pay += _money(outstanding)
        elif entry.get("entry_type") in PAYMENT_TYPES:
            on_account += _money(outstanding)
    return {
        "balance": float(balance),
        "to_receive": float(to_receive),
        "to_pay": float(to_pay),
        "on_account": float(on_account),
        "open_count": open_count,
    }


def plan_allocations(
    charges: list[dict[str, Any]], amount: float
) -> list[dict[str, Any]]:
    """Spread `amount` across open charges, oldest first.

    `charges` must already be ordered and carry an `outstanding` field. Returns
    `[{"charge_id": ..., "amount": ...}]`. Any remainder is left unallocated
    (the payment sits on account).
    """
    remaining = _money(amount)
    out: list[dict[str, Any]] = []
    for charge in charges:
        if remaining <= Decimal("0"):
            break
        outstanding = _money(charge.get("outstanding") or 0)
        if outstanding <= Decimal("0"):
            continue
        take = min(remaining, outstanding)
        if take <= Decimal("0"):
            continue
        out.append({"charge_id": str(charge.get("id") or ""), "amount": float(take)})
        remaining -= take
    return out


# ------------------------------------------------------------------ data loading


def _fetch_entries(
    *, account_type: str | None = None, account_id: str | None = None
) -> list[dict[str, Any]]:
    params: dict[str, str] = {
        "select": "*",
        "order": "entry_date.asc,created_at.asc",
        "limit": "20000",
    }
    if account_type and account_id:
        params[_account_column(account_type)] = f"eq.{account_id}"
    elif account_type:
        params[_account_column(account_type)] = "not.is.null"
    response = _http().get(
        f"{_supabase_url()}/rest/v1/{LEDGER_TABLE}",
        headers=_service_headers(),
        params=params,
        timeout=HTTP_TIMEOUT,
    )
    if response.status_code != 200:
        raise RuntimeError(f"List ledger failed ({response.status_code}): {response.text[:300]}")
    rows = response.json()
    return rows if isinstance(rows, list) else []


def _fetch_allocations(charge_ids: list[str] | None = None) -> list[dict[str, Any]]:
    params: dict[str, str] = {"select": "*", "limit": "20000"}
    if charge_ids is not None:
        if not charge_ids:
            return []
        params["charge_id"] = "in.(" + ",".join(charge_ids) + ")"
    response = _http().get(
        f"{_supabase_url()}/rest/v1/{ALLOCATIONS_TABLE}",
        headers=_service_headers(),
        params=params,
        timeout=HTTP_TIMEOUT,
    )
    if response.status_code != 200:
        raise RuntimeError(
            f"List ledger allocations failed ({response.status_code}): {response.text[:300]}"
        )
    rows = response.json()
    return rows if isinstance(rows, list) else []


def _allocated_by_entry(allocations: list[dict[str, Any]]) -> dict[str, float]:
    """Charge id → amount cleared, and payment id → amount applied."""
    totals: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for row in allocations:
        amount = _money(row.get("amount"))
        totals[str(row.get("charge_id") or "")] += amount
        totals[str(row.get("payment_id") or "")] += amount
    return {k: float(v) for k, v in totals.items()}


def _account_names(account_type: str) -> dict[str, dict[str, Any]]:
    response = _http().get(
        f"{_supabase_url()}/rest/v1/{_account_table(account_type)}",
        headers=_service_headers(),
        params={"select": "id,name,status,is_archived", "order": "name.asc", "limit": "5000"},
        timeout=HTTP_TIMEOUT,
    )
    if response.status_code != 200:
        return {}
    return {str(r["id"]): r for r in (response.json() or []) if r.get("id")}


def _ipo_names() -> dict[str, dict[str, Any]]:
    response = _http().get(
        f"{_supabase_url()}/rest/v1/{IPO_TABLE}",
        headers=_service_headers(),
        params={"select": "id,name,display_name,listing_date", "limit": "5000"},
        timeout=HTTP_TIMEOUT,
    )
    if response.status_code != 200:
        return {}
    return {str(r["id"]): r for r in (response.json() or []) if r.get("id")}


# --------------------------------------------------------------------- overview


def ledger_overview(side: str = "buyer", *, include_settled: bool = False) -> dict[str, Any]:
    """Every account on one side with its outstanding — the landing view."""
    if (side or "").strip().casefold() in ("all", "everyone", "both"):
        accounts = _combined_accounts(include_settled=include_settled)
        return {
            "side": "all",
            "account_type": "all",
            "accounts": accounts,
            "totals": {
                "to_receive": float(_money(sum(a["to_receive"] for a in accounts))),
                "to_pay": float(_money(sum(a["to_pay"] for a in accounts))),
                "balance": float(_money(sum(a["balance"] for a in accounts))),
                "open_count": sum(int(a["open_count"]) for a in accounts),
                "account_count": len(accounts),
            },
        }
    account_type = side_to_account_type(side)
    names = _account_names(account_type)
    entries = _fetch_entries(account_type=account_type)
    allocated = _allocated_by_entry(_fetch_allocations())
    column = _account_column(account_type)

    by_account: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for entry in entries:
        key = str(entry.get(column) or "")
        if key:
            by_account[key].append(entry)

    accounts: list[dict[str, Any]] = []
    for account_id, meta in names.items():
        rows = by_account.get(account_id, [])
        summary = summarize_entries(rows, allocated)
        last_activity = max(
            (str(r.get("entry_date") or "")[:10] for r in rows), default=""
        )
        record = {
            "account_type": account_type,
            "account_id": account_id,
            "side": "seller" if account_type == ACCOUNT_SELL_PARTY else "client",
            "name": meta.get("name") or "—",
            "status": meta.get("status") or "Active",
            "is_archived": bool(meta.get("is_archived", False)),
            "entry_count": len(rows),
            "last_activity": last_activity,
            "direction": balance_direction(float(summary["balance"])),
            **summary,
        }
        if not include_settled and record["open_count"] == 0 and abs(record["balance"]) < EPSILON:
            continue
        accounts.append(record)

    # Biggest pending first, then by name.
    accounts.sort(key=lambda a: (-abs(float(a["balance"])), (a["name"] or "").lower()))
    totals = {
        "to_receive": float(_money(sum(a["to_receive"] for a in accounts))),
        "to_pay": float(_money(sum(a["to_pay"] for a in accounts))),
        "balance": float(_money(sum(a["balance"] for a in accounts))),
        "open_count": sum(int(a["open_count"]) for a in accounts),
        "account_count": len(accounts),
    }
    return {"side": side, "account_type": account_type, "accounts": accounts, "totals": totals}


def _combined_accounts(*, include_settled: bool = False) -> list[dict[str, Any]]:
    """Clients and sell parties in one list, each tagged with its side."""
    rows: list[dict[str, Any]] = []
    for account_type in ACCOUNT_TYPES:
        rows.extend(ledger_overview(account_type, include_settled=include_settled)["accounts"])
    rows.sort(key=lambda a: (-abs(float(a["balance"])), (a["name"] or "").lower()))
    return rows


def ledger_by_ipo(ipo_id: str) -> dict[str, Any]:
    """Every client AND sell party on one IPO — the Settlement & Reports layout,
    plus what has actually been paid since."""
    entries = [
        e for e in _fetch_entries() if str(e.get("ipo_id") or "") == str(ipo_id)
    ]
    allocated = _allocated_by_entry(_fetch_allocations())
    names = {
        ACCOUNT_PARTY: _account_names(ACCOUNT_PARTY),
        ACCOUNT_SELL_PARTY: _account_names(ACCOUNT_SELL_PARTY),
    }
    ipo = _ipo_names().get(str(ipo_id)) or {}

    rows: list[dict[str, Any]] = []
    for entry in entries:
        if entry.get("entry_type") not in CHARGE_TYPES:
            continue
        account_type = (
            ACCOUNT_SELL_PARTY if entry.get("sell_party_id") else ACCOUNT_PARTY
        )
        account_id = str(entry.get(_account_column(account_type)) or "")
        meta = names[account_type].get(account_id) or {}
        amount = float(entry.get("amount") or 0)
        paid = float(allocated.get(str(entry.get("id") or ""), 0.0))
        rows.append(
            ledger_entry_to_json(entry)
            | {
                "account_type": account_type,
                "account_id": account_id,
                "name": meta.get("name") or "—",
                "side": "seller" if account_type == ACCOUNT_SELL_PARTY else "client",
                "paid": paid,
                "outstanding": _outstanding(amount, paid),
                "status": charge_status(amount, paid),
                "direction": charge_direction(amount),
            }
        )
    rows.sort(key=lambda r: (r["side"] != "seller", -abs(r["amount"])))

    sellers = [r for r in rows if r["side"] == "seller"]
    clients = [r for r in rows if r["side"] == "client"]
    collect = sum(r["outstanding"] for r in rows if r["amount"] > 0)
    give = sum(r["outstanding"] for r in rows if r["amount"] < 0)
    return {
        "ipo": {
            "id": str(ipo_id),
            "name": ipo.get("display_name") or ipo.get("name") or "—",
            "listing_date": str(ipo.get("listing_date") or "")[:10],
        },
        "rows": rows,
        "totals": {
            "charged_sellers": float(_money(sum(r["amount"] for r in sellers))),
            "charged_clients": float(_money(sum(r["amount"] for r in clients))),
            "collect": float(_money(collect)),
            "give": float(_money(give)),
            "net": float(_money(collect - give)),
            "brokerage": float(_money(sum(r["amount"] for r in rows))),
            "open_count": sum(1 for r in rows if r["outstanding"] > EPSILON),
            "row_count": len(rows),
        },
    }


def ipos_with_ledger() -> list[dict[str, Any]]:
    """IPO picker for the by-IPO view — newest listing first."""
    entries = _fetch_entries()
    allocated = _allocated_by_entry(_fetch_allocations())
    ipos = _ipo_names()

    by_ipo: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for entry in entries:
        key = str(entry.get("ipo_id") or "")
        if key:
            by_ipo[key].append(entry)

    out: list[dict[str, Any]] = []
    for ipo_id, rows in by_ipo.items():
        meta = ipos.get(ipo_id) or {}
        charges = [r for r in rows if r.get("entry_type") in CHARGE_TYPES]
        collect = Decimal("0")
        give = Decimal("0")
        open_count = 0
        for charge in charges:
            amount = float(charge.get("amount") or 0)
            paid = float(allocated.get(str(charge.get("id") or ""), 0.0))
            outstanding = _outstanding(amount, paid)
            if outstanding > EPSILON:
                open_count += 1
                if amount > 0:
                    collect += _money(outstanding)
                else:
                    give += _money(outstanding)
        out.append(
            {
                "ipo_id": ipo_id,
                "name": meta.get("display_name") or meta.get("name") or "—",
                "listing_date": str(meta.get("listing_date") or "")[:10],
                "party_count": len(charges),
                "open_count": open_count,
                "collect": float(collect),
                "give": float(give),
                "net": float(collect - give),
            }
        )
    out.sort(key=lambda i: (i["listing_date"] or "", i["name"]), reverse=True)
    return out


def ledger_summary() -> dict[str, Any]:
    """The middleman picture: take from sell parties → give to clients → keep the rest.

    Both sides in one payload, so the page never shows a half-truth like
    "to receive 0" just because the Clients tab happens to be open.
    """
    buyers = ledger_overview(ACCOUNT_PARTY)["totals"]
    sellers = ledger_overview(ACCOUNT_SELL_PARTY)["totals"]

    collect = float(_money(buyers["to_receive"] + sellers["to_receive"]))
    give = float(_money(buyers["to_pay"] + sellers["to_pay"]))
    return {
        "buyers": buyers,
        "sellers": sellers,
        # Headline flow, in the order the money actually moves.
        "collect_from_sellers": sellers["to_receive"],
        "collect_from_clients": buyers["to_receive"],
        "give_to_clients": buyers["to_pay"],
        "give_to_sellers": sellers["to_pay"],
        "collect_total": collect,
        "give_total": give,
        # What is left over once everything is collected and paid out.
        "net": float(_money(collect - give)),
        "open_count": int(buyers["open_count"]) + int(sellers["open_count"]),
    }


# ---------------------------------------------------------------- single account


def account_ledger(account_type: str, account_id: str) -> dict[str, Any]:
    """One person's ledger: IPO-wise open items + full history."""
    column = _account_column(account_type)
    entries = _fetch_entries(account_type=account_type, account_id=account_id)
    allocated = _allocated_by_entry(_fetch_allocations())
    ipos = _ipo_names()
    names = _account_names(account_type)
    meta = names.get(account_id) or {}

    open_items: list[dict[str, Any]] = []
    history: list[dict[str, Any]] = []
    running = Decimal("0")
    for row in entries:
        amount = float(row.get("amount") or 0)
        running += _money(amount)
        entry_id = str(row.get("id") or "")
        paid = float(allocated.get(entry_id, 0.0))
        outstanding = _outstanding(amount, paid)
        ipo = ipos.get(str(row.get("ipo_id") or "")) or {}
        ipo_name = ipo.get("display_name") or ipo.get("name") or ""
        item = ledger_entry_to_json(row) | {
            "account_type": account_type,
            "account_id": account_id,
            "ipo_name": ipo_name,
            "listing_date": str(ipo.get("listing_date") or "")[:10],
            "paid": paid,
            "outstanding": outstanding,
            "running_balance": float(running),
            "is_charge": row.get("entry_type") in CHARGE_TYPES,
            "status": charge_status(amount, paid),
            "direction": charge_direction(amount),
        }
        history.append(item)
        if item["is_charge"]:
            open_items.append(item)

    summary = summarize_entries(entries, allocated)
    return {
        "account": {
            "account_type": account_type,
            "account_id": account_id,
            "name": meta.get("name") or "—",
            "side": "seller" if account_type == ACCOUNT_SELL_PARTY else "buyer",
        },
        # Back-compat with the old single-party shape.
        "party_id": account_id if account_type == ACCOUNT_PARTY else None,
        "direction": balance_direction(float(summary["balance"])),
        "open_items": open_items,
        "entries": history,
        **summary,
    }


def list_ledger(party_id: str) -> dict[str, Any]:
    """Back-compat wrapper — buy party only."""
    return account_ledger(ACCOUNT_PARTY, party_id)


def account_balance(account_type: str, account_id: str) -> float:
    entries = _fetch_entries(account_type=account_type, account_id=account_id)
    total = Decimal("0")
    for row in entries:
        total += _money(row.get("amount"))
    return float(total)


def party_balance(party_id: str) -> float:
    return account_balance(ACCOUNT_PARTY, party_id)


# ----------------------------------------------------------------- writing entries


def _append_entry(
    *,
    account_type: str = ACCOUNT_PARTY,
    account_id: str,
    amount: float,
    entry_type: str,
    entry_date: str,
    notes: str = "",
    ipo_id: str | None = None,
    reference_type: str = "",
    reference_id: str | None = None,
) -> dict[str, Any]:
    current = account_balance(account_type, account_id)
    new_balance = float(_money(current) + _money(amount))
    payload = {
        _account_column(account_type): account_id,
        "ipo_id": ipo_id,
        "entry_type": entry_type,
        "amount": float(_money(amount)),
        "balance_after": new_balance,
        "reference_type": reference_type,
        "reference_id": reference_id,
        "entry_date": entry_date,
        "notes": notes or "",
        "updated_at": _now(),
    }
    response = _http().post(
        f"{_supabase_url()}/rest/v1/{LEDGER_TABLE}",
        headers=_service_headers(
            {"Content-Type": "application/json", "Prefer": "return=representation"}
        ),
        json=payload,
        timeout=HTTP_TIMEOUT,
    )
    if response.status_code not in (200, 201):
        raise RuntimeError(
            f"Create ledger entry failed ({response.status_code}): {response.text[:300]}"
        )
    data = response.json()
    row = data[0] if isinstance(data, list) else data
    return ledger_entry_to_json(row)


def _delete_entry(entry_id: str) -> None:
    _http().delete(
        f"{_supabase_url()}/rest/v1/{LEDGER_TABLE}",
        headers=_service_headers({"Prefer": "return=minimal"}),
        params={"id": f"eq.{entry_id}"},
        timeout=HTTP_TIMEOUT,
    )


def _insert_allocations(payment_id: str, allocations: list[dict[str, Any]]) -> None:
    rows = [
        {
            "payment_id": payment_id,
            "charge_id": a["charge_id"],
            "amount": float(_money(a["amount"])),
        }
        for a in allocations
        if float(a.get("amount") or 0) > 0
    ]
    if not rows:
        return
    response = _http().post(
        f"{_supabase_url()}/rest/v1/{ALLOCATIONS_TABLE}",
        headers=_service_headers(
            {"Content-Type": "application/json", "Prefer": "return=minimal"}
        ),
        json=rows,
        timeout=HTTP_TIMEOUT,
    )
    if response.status_code not in (200, 201):
        raise RuntimeError(
            f"Allocate payment failed ({response.status_code}): {response.text[:300]}"
        )


def post_settlement_entries(
    *,
    ipo_id: str,
    settlement_id: str,
    party_nets: dict[str, float],
    sell_party_nets: dict[str, float] | None = None,
) -> list[dict[str, Any]]:
    """One Settlement charge per account. amount = + receivable / − payable."""
    today = datetime.now(timezone.utc).date().isoformat()
    created: list[dict[str, Any]] = []
    buckets: list[tuple[str, dict[str, float]]] = [(ACCOUNT_PARTY, party_nets or {})]
    if sell_party_nets:
        buckets.append((ACCOUNT_SELL_PARTY, sell_party_nets))

    for account_type, nets in buckets:
        column = _account_column(account_type)
        for account_id, net in nets.items():
            if not account_id or abs(float(net)) < 1e-12:
                continue
            # Skip if already posted (the unique index would reject it anyway).
            check = _http().get(
                f"{_supabase_url()}/rest/v1/{LEDGER_TABLE}",
                headers=_service_headers(),
                params={
                    column: f"eq.{account_id}",
                    "reference_id": f"eq.{settlement_id}",
                    "entry_type": "eq.Settlement",
                    "select": "id",
                    "limit": "1",
                },
                timeout=HTTP_TIMEOUT,
            )
            if check.status_code == 200 and isinstance(check.json(), list) and check.json():
                continue
            created.append(
                _append_entry(
                    account_type=account_type,
                    account_id=account_id,
                    amount=float(net),
                    entry_type="Settlement",
                    entry_date=today,
                    notes="IPO settlement finalized",
                    ipo_id=ipo_id,
                    reference_type="settlement",
                    reference_id=settlement_id,
                )
            )
    return created


def settlement_revision_deltas(
    *,
    posted: dict[str, float],
    desired: dict[str, float],
) -> dict[str, float]:
    """Pure: Adjustment amount per account so posted + delta = desired.

    Accounts with an unchanged net are omitted. Accounts that disappear from
    the new settlement get a full reversing delta (−posted).
    """
    keys = set(posted) | set(desired)
    out: dict[str, float] = {}
    for key in keys:
        old = float(posted.get(key) or 0)
        new = float(desired.get(key) or 0)
        delta = float(_money(new) - _money(old))
        if abs(delta) < EPSILON:
            continue
        out[key] = delta
    return out


def _posted_settlement_nets_for_ipo(ipo_id: str) -> tuple[dict[str, float], dict[str, float]]:
    """Sum of Settlement + settlement_revision Adjustments already on the ledger for this IPO.

    Manual adjustments (reference_type=manual) are intentionally excluded so
    business-logic revisions never clobber write-offs or opening balances.
    """
    response = _http().get(
        f"{_supabase_url()}/rest/v1/{LEDGER_TABLE}",
        headers=_service_headers(),
        params={
            "ipo_id": f"eq.{ipo_id}",
            "select": "party_id,sell_party_id,entry_type,reference_type,amount",
            "limit": "20000",
        },
        timeout=HTTP_TIMEOUT,
    )
    if response.status_code != 200:
        raise RuntimeError(
            f"Load posted settlement nets failed ({response.status_code}): {response.text[:300]}"
        )
    party: dict[str, float] = defaultdict(float)
    sell: dict[str, float] = defaultdict(float)
    for row in response.json() or []:
        entry_type = row.get("entry_type") or ""
        ref = row.get("reference_type") or ""
        keep = (entry_type == "Settlement" and ref == "settlement") or (
            entry_type == "Adjustment" and ref == "settlement_revision"
        )
        if not keep:
            continue
        amount = float(row.get("amount") or 0)
        if row.get("sell_party_id"):
            sell[str(row["sell_party_id"])] += amount
        elif row.get("party_id"):
            party[str(row["party_id"])] += amount
    return (
        {k: float(_money(v)) for k, v in party.items()},
        {k: float(_money(v)) for k, v in sell.items()},
    )


def post_settlement_revisions(
    *,
    ipo_id: str,
    settlement_id: str,
    party_nets: dict[str, float],
    sell_party_nets: dict[str, float] | None = None,
) -> list[dict[str, Any]]:
    """Append-only corrections so ledger totals match the latest settlement math.

    Never overwrites Settlement / Payment / manual Adjustment rows. Posts one
    Adjustment (reference_type=settlement_revision) per account whose desired
    net differs from what is already posted for this IPO.
    """
    today = datetime.now(timezone.utc).date().isoformat()
    posted_party, posted_sell = _posted_settlement_nets_for_ipo(ipo_id)
    party_deltas = settlement_revision_deltas(posted=posted_party, desired=party_nets or {})
    sell_deltas = settlement_revision_deltas(
        posted=posted_sell, desired=sell_party_nets or {}
    )
    created: list[dict[str, Any]] = []
    for account_type, deltas in (
        (ACCOUNT_PARTY, party_deltas),
        (ACCOUNT_SELL_PARTY, sell_deltas),
    ):
        for account_id, delta in deltas.items():
            posted_map = posted_party if account_type == ACCOUNT_PARTY else posted_sell
            desired_map = (
                (party_nets or {})
                if account_type == ACCOUNT_PARTY
                else (sell_party_nets or {})
            )
            posted = posted_map.get(account_id, 0.0)
            desired = desired_map.get(account_id, 0.0)
            created.append(
                _append_entry(
                    account_type=account_type,
                    account_id=account_id,
                    amount=float(delta),
                    entry_type="Adjustment",
                    entry_date=today,
                    notes=(
                        f"Settlement revision: {float(posted):,.4f} → {float(desired):,.4f}"
                    ),
                    ipo_id=ipo_id,
                    reference_type="settlement_revision",
                    reference_id=settlement_id,
                )
            )
    return created


def pay_open_items(
    *,
    account_type: str,
    account_id: str,
    charge_ids: list[str],
    entry_date: str,
    amount: float | None = None,
    notes: str = "",
) -> dict[str, Any]:
    """Tick some IPO charges, pay them off with a single entry.

    `amount` defaults to the full outstanding of the ticked charges; a smaller
    amount is spread oldest-first. All ticked charges must move money the same
    way — you cannot mix "to receive" and "to pay" in one payment.
    """
    if not charge_ids:
        raise ValueError("Select at least one IPO to settle.")
    data = account_ledger(account_type, account_id)
    wanted = set(charge_ids)
    charges = [
        item
        for item in data["open_items"]
        if item["id"] in wanted and item["outstanding"] > EPSILON
    ]
    if not charges:
        raise ValueError("Those entries are already settled.")

    directions = {"receive" if float(c["amount"]) > 0 else "pay" for c in charges}
    if len(directions) > 1:
        raise ValueError(
            "Selected IPOs mix money coming in and going out. "
            "Settle the 'to receive' ones separately from the 'to pay' ones."
        )
    direction = directions.pop()

    total = float(_money(sum(c["outstanding"] for c in charges)))
    pay_amount = total if amount is None else float(_money(amount))
    if pay_amount <= 0:
        raise ValueError("Amount must be greater than 0.")
    if pay_amount - total > EPSILON:
        raise ValueError(
            f"Amount {pay_amount:,.2f} is more than the selected outstanding {total:,.2f}."
        )

    entry_type = "PaymentReceived" if direction == "receive" else "PaymentPaid"
    # PaymentReceived reduces a receivable (+) → negative to balance.
    signed = -pay_amount if entry_type == "PaymentReceived" else pay_amount
    ipo_id = charges[0].get("ipo_id") if len(charges) == 1 else None

    payment = _append_entry(
        account_type=account_type,
        account_id=account_id,
        amount=signed,
        entry_type=entry_type,
        entry_date=entry_date,
        notes=notes,
        ipo_id=ipo_id,
        reference_type="payment",
    )
    try:
        _insert_allocations(payment["id"], plan_allocations(charges, pay_amount))
    except Exception:
        # Never leave an unallocated payment behind from a half-failed write.
        _delete_entry(payment["id"])
        raise

    return account_ledger(account_type, account_id) | {"last_entry": payment}


def pay_open_items_batch(
    *, payments: list[dict[str, Any]], entry_date: str, notes: str = ""
) -> dict[str, Any]:
    """Settle several accounts in one go (overview checkboxes).

    Each item is `{account_type, account_id, charge_ids?, amount?}`. Omitting
    `charge_ids` settles every open item on that account in one direction, which
    is only valid when the account has no mixed directions.
    """
    if not payments:
        raise ValueError("Select at least one account to settle.")
    done: list[dict[str, Any]] = []
    failed: list[dict[str, Any]] = []
    for item in payments:
        account_type = item.get("account_type") or ACCOUNT_PARTY
        account_id = str(item.get("account_id") or "")
        try:
            charge_ids = item.get("charge_ids")
            if not charge_ids:
                ledger = account_ledger(account_type, account_id)
                charge_ids = [
                    c["id"] for c in ledger["open_items"] if c["outstanding"] > EPSILON
                ]
            result = pay_open_items(
                account_type=account_type,
                account_id=account_id,
                charge_ids=list(charge_ids),
                entry_date=entry_date,
                amount=item.get("amount"),
                notes=item.get("notes") or notes,
            )
            done.append(
                {
                    "account_id": account_id,
                    "account_type": account_type,
                    "name": result["account"]["name"],
                    "amount": abs(float(result["last_entry"]["amount"])),
                }
            )
        except Exception as exc:
            failed.append({"account_id": account_id, "error": str(exc)})
    return {"settled": done, "failed": failed}


def record_payment(
    *,
    party_id: str | None = None,
    entry_type: str,
    amount: float,
    entry_date: str,
    notes: str = "",
    ipo_id: str | None = None,
    account_type: str = ACCOUNT_PARTY,
    account_id: str | None = None,
    auto_allocate: bool = True,
) -> dict[str, Any]:
    """Record an unlinked payment. By default it auto-clears the oldest charges."""
    if entry_type not in PAYMENT_TYPES:
        raise ValueError("entry_type must be PaymentReceived or PaymentPaid")
    if amount <= 0:
        raise ValueError("amount must be > 0")
    target = account_id or party_id
    if not target:
        raise ValueError("account_id is required")

    signed = -float(amount) if entry_type == "PaymentReceived" else float(amount)
    entry = _append_entry(
        account_type=account_type,
        account_id=target,
        amount=signed,
        entry_type=entry_type,
        entry_date=entry_date,
        notes=notes,
        ipo_id=ipo_id,
        reference_type="payment",
    )
    if auto_allocate:
        data = account_ledger(account_type, target)
        want_receivable = entry_type == "PaymentReceived"
        charges = [
            c
            for c in data["open_items"]
            if c["outstanding"] > EPSILON and (float(c["amount"]) > 0) == want_receivable
        ]
        try:
            _insert_allocations(entry["id"], plan_allocations(charges, float(amount)))
        except Exception:
            _delete_entry(entry["id"])
            raise
    return account_ledger(account_type, target) | {"last_entry": entry}


# ------------------------------------------------------------------ editing

# Settlement charges are owned by the settlement — editing them here would
# silently desync the Ledger from Settlement & Reports. Everything else is
# operator data and is fair game.
EDITABLE_TYPES = ("Adjustment", "PaymentReceived", "PaymentPaid")


def _get_entry(entry_id: str) -> dict[str, Any]:
    response = _http().get(
        f"{_supabase_url()}/rest/v1/{LEDGER_TABLE}",
        headers=_service_headers(),
        params={"id": f"eq.{entry_id}", "select": "*", "limit": "1"},
        timeout=HTTP_TIMEOUT,
    )
    if response.status_code != 200:
        raise RuntimeError(f"Get ledger entry failed ({response.status_code})")
    rows = response.json()
    if not isinstance(rows, list) or not rows:
        raise LookupError("Ledger entry not found.")
    return rows[0]


def _entry_account(row: dict[str, Any]) -> tuple[str, str]:
    if row.get("sell_party_id"):
        return ACCOUNT_SELL_PARTY, str(row["sell_party_id"])
    return ACCOUNT_PARTY, str(row.get("party_id") or "")


def _delete_allocations_for_payment(payment_id: str) -> None:
    _http().delete(
        f"{_supabase_url()}/rest/v1/{ALLOCATIONS_TABLE}",
        headers=_service_headers({"Prefer": "return=minimal"}),
        params={"payment_id": f"eq.{payment_id}"},
        timeout=HTTP_TIMEOUT,
    )


def add_adjustment(
    *,
    account_type: str,
    account_id: str,
    amount: float,
    entry_date: str,
    notes: str = "",
    ipo_id: str | None = None,
) -> dict[str, Any]:
    """Manual charge: opening balance carried forward, write-off, discount, correction.

    `amount` is signed the same way as everything else — **positive** means
    collect more from them, **negative** means pay them more. It becomes an
    open item, so it can be part-paid and settled like any IPO charge.
    """
    if abs(float(amount)) < EPSILON:
        raise ValueError("Amount must not be zero.")
    if not (notes or "").strip():
        raise ValueError("Write a short reason — this is a manual entry.")
    _append_entry(
        account_type=account_type,
        account_id=account_id,
        amount=float(amount),
        entry_type="Adjustment",
        entry_date=entry_date,
        notes=notes,
        ipo_id=ipo_id,
        reference_type="manual",
    )
    return account_ledger(account_type, account_id)


def update_ledger_entry(
    entry_id: str,
    *,
    entry_date: str | None = None,
    notes: str | None = None,
    amount: float | None = None,
) -> dict[str, Any]:
    """Correct a payment or adjustment.

    Date and notes are always editable. Changing a payment's amount re-spreads
    it across the open charges; reducing an adjustment below what has already
    been paid against it is refused rather than silently leaving a broken link.
    """
    row = _get_entry(entry_id)
    entry_type = row.get("entry_type") or ""
    if entry_type not in EDITABLE_TYPES:
        raise ValueError(
            "Settlement entries cannot be edited here — they belong to the "
            "settlement. Add an Adjustment instead, so the change is visible."
        )
    account_type, account_id = _entry_account(row)

    patch: dict[str, Any] = {"updated_at": _now()}
    if entry_date:
        patch["entry_date"] = entry_date
    if notes is not None:
        patch["notes"] = notes

    is_payment = entry_type in PAYMENT_TYPES
    new_amount: float | None = None
    if amount is not None:
        magnitude = abs(float(amount))
        if magnitude < EPSILON:
            raise ValueError("Amount must not be zero.")
        if is_payment:
            # Keep the payment's direction; only the size is editable.
            new_amount = -magnitude if entry_type == "PaymentReceived" else magnitude
        else:
            allocated = _allocated_by_entry(_fetch_allocations([entry_id])).get(entry_id, 0.0)
            if magnitude + EPSILON < allocated:
                raise ValueError(
                    f"{magnitude:,.2f} is less than the {allocated:,.2f} already "
                    "paid against this entry. Undo the payment first."
                )
            new_amount = float(amount)
        patch["amount"] = float(_money(new_amount))

    response = _http().patch(
        f"{_supabase_url()}/rest/v1/{LEDGER_TABLE}",
        headers=_service_headers({"Content-Type": "application/json"}),
        params={"id": f"eq.{entry_id}"},
        json=patch,
        timeout=HTTP_TIMEOUT,
    )
    if response.status_code not in (200, 204):
        raise RuntimeError(
            f"Update ledger entry failed ({response.status_code}): {response.text[:300]}"
        )

    # A resized payment must be re-spread, or it would still claim the old amounts.
    if is_payment and new_amount is not None:
        _delete_allocations_for_payment(entry_id)
        data = account_ledger(account_type, account_id)
        want_receivable = entry_type == "PaymentReceived"
        charges = [
            c
            for c in data["open_items"]
            if c["outstanding"] > EPSILON and (float(c["amount"]) > 0) == want_receivable
        ]
        _insert_allocations(entry_id, plan_allocations(charges, abs(new_amount)))

    return account_ledger(account_type, account_id)


def delete_ledger_entry(entry_id: str) -> dict[str, Any]:
    """Undo a payment or adjustment. Settlement charges are never deleted here."""
    row = _get_entry(entry_id)
    entry_type = row.get("entry_type") or ""
    if entry_type not in EDITABLE_TYPES:
        raise ValueError(
            "Settlement entries cannot be deleted here — un-finalize the "
            "settlement instead, so both pages stay in step."
        )
    account_type, account_id = _entry_account(row)
    # Allocations cascade on delete, which re-opens whatever this payment closed.
    _delete_entry(entry_id)
    return account_ledger(account_type, account_id)


# ------------------------------------------------------------------------ export


def build_statement_excel(
    account_id: str, account_name: str = "", account_type: str = ACCOUNT_PARTY
) -> bytes:
    data = account_ledger(account_type, account_id)
    wb = Workbook()

    ws = wb.active
    ws.title = "Outstanding"
    ws.append(["Account", account_name or data["account"]["name"]])
    ws.append(["Side", "Sell party" if account_type == ACCOUNT_SELL_PARTY else "Client"])
    ws.append(["Net outstanding", data["balance"], data["direction"]])
    ws.append(["Collect", data["to_receive"]])
    ws.append(["Pay them", data["to_pay"]])
    ws.append(["Advance (not linked to IPO yet)", data["on_account"]])
    ws.append([])
    ws.append(["IPO", "Date", "Amount", "Paid", "Outstanding", "Status", "Direction"])
    for item in data["open_items"]:
        ws.append(
            [
                item.get("ipo_name") or "—",
                item.get("entry_date"),
                item.get("amount"),
                item.get("paid"),
                item.get("outstanding"),
                item.get("status"),
                item.get("direction"),
            ]
        )

    ws2 = wb.create_sheet("All entries")
    ws2.append(["Date", "Type", "IPO", "Amount", "Paid", "Outstanding", "Balance", "Notes"])
    for e in data["entries"]:
        ws2.append(
            [
                e.get("entry_date"),
                e.get("entry_type"),
                e.get("ipo_name") or "",
                e.get("amount"),
                e.get("paid"),
                e.get("outstanding"),
                e.get("running_balance"),
                e.get("notes"),
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_overview_excel(side: str = "buyer", *, include_settled: bool = False) -> bytes:
    data = ledger_overview(side, include_settled=include_settled)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sellers" if data["account_type"] == ACCOUNT_SELL_PARTY else "Clients"
    ws.append(["Name", "Pending IPOs", "Collect", "Pay them", "Net pending", "Direction", "Last activity"])
    for a in data["accounts"]:
        ws.append(
            [
                a.get("name"),
                a.get("open_count"),
                a.get("to_receive"),
                a.get("to_pay"),
                a.get("balance"),
                a.get("direction"),
                a.get("last_activity"),
            ]
        )
    totals = data["totals"]
    ws.append([])
    ws.append(["TOTAL", totals["open_count"], totals["to_receive"], totals["to_pay"], totals["balance"]])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
